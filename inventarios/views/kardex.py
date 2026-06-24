from django.shortcuts import render, get_object_or_404

from catalogos.models import Almacen, Producto, Categoria

from ..models import EntradaInventarioDetalle, InventarioStock, SalidaInventarioDetalle

from ..utils import get_almacen_default

from datetime import datetime

from decimal import Decimal

from django.db.models import Q

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from accounts.decorators import grupos_requeridos, permiso_requerido

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from io import BytesIO

@permiso_requerido("inventarios.view_inventariostock")
def inventario_actual(request):
    almacenes_qs = Almacen.objects.filter(es_activo=True).order_by("tipo", "nombre")
    almacen = get_almacen_default()
    almacen_id = (request.GET.get("almacen") or "").strip()
    if almacen_id.isdigit():
        almacen = almacenes_qs.filter(pk=int(almacen_id)).first() or almacen

    q = (request.GET.get("q") or "").strip()
    categoria_id = (request.GET.get("categoria") or "").strip()
    solo_bajos = request.GET.get("solo_bajos") == "on"

    productos = Producto.objects.select_related("categoria").all()

    if q:
        productos = productos.filter(
            Q(nombre__icontains=q) |
            Q(clave_sat__icontains=q) |
            Q(metrica__icontains=q)
        )

    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)

    stock_map = {}
    if almacen:
        stock_map = {
            s.producto_id: (s.cantidad or 0)
            for s in InventarioStock.objects.filter(almacen_id=almacen.id)
        }

    categorias = Categoria.objects.all().order_by("nombre")

    productos = list(productos.order_by("nombre"))
    for p in productos:
        p.stock_almacen = stock_map.get(p.id, 0)

    if solo_bajos:
        productos = [p for p in productos if p.stock_minimo is not None and p.stock_almacen <= p.stock_minimo]

    return render(request, "inventarios/inventario_actual.html", {
        "productos": productos,
        "categorias": categorias,
        "almacenes": almacenes_qs,
        "almacen": almacen,
        "almacen_id": almacen_id,
        "q": q,
        "categoria_id": categoria_id,
        "solo_bajos": solo_bajos,
    }) 

@permiso_requerido("inventarios.view_inventariostock")
def kardex(request):
    producto_id = (request.GET.get("producto") or "").strip()
    almacenes_qs = Almacen.objects.filter(es_activo=True).order_by("tipo", "nombre")
    almacen = get_almacen_default()
    almacen_id = (request.GET.get("almacen") or "").strip()
    if almacen_id.isdigit():
        almacen = almacenes_qs.filter(pk=int(almacen_id)).first() or almacen

    fecha_ini = (request.GET.get("fecha_ini") or "").strip()
    fecha_fin = (request.GET.get("fecha_fin") or "").strip()

    productos = Producto.objects.all().order_by("nombre")

    producto = None
    movimientos = []
    saldo_inicial = None
    saldo_final = None
    tot_entradas = None
    tot_salidas = None

    # Si no hay producto seleccionado, sólo mostramos el selector
    if not producto_id:
        return render(request, "inventarios/kardex.html", {
            "productos": productos,
            "almacenes": almacenes_qs,
            "almacen": almacen,
            "almacen_id": almacen_id,
            "producto": None,
            "producto_id": "",
            "fecha_ini": fecha_ini,
            "fecha_fin": fecha_fin,
            "movimientos": [],
        })

    producto = get_object_or_404(Producto, pk=producto_id)

    # Parse fechas (si vienen)
    dt_ini = None
    dt_fin = None
    try:
        if fecha_ini:
            dt_ini = datetime.strptime(fecha_ini, "%Y-%m-%d").date()
        if fecha_fin:
            dt_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    except ValueError:
        # Si llega mal el formato, lo ignoramos (podrías mostrar mensaje)
        dt_ini = None
        dt_fin = None

    # ---------- ENTRADAS ----------
    q_entradas = (
        EntradaInventarioDetalle.objects
        .select_related("entrada", "almacen")
        .filter(producto_id=producto.id)
    )

    if almacen:                
        q_entradas = q_entradas.filter(
            Q(almacen_id=almacen.id) |
            Q(almacen__isnull=True, entrada__almacen_id=almacen.id)
        )

    if dt_ini:
        q_entradas = q_entradas.filter(entrada__fecha__gte=dt_ini)

    if dt_fin:
        q_entradas = q_entradas.filter(entrada__fecha__lte=dt_fin)


    # ---------- SALIDAS ----------
    q_salidas = (
        SalidaInventarioDetalle.objects
        .select_related("salida")
        .filter(producto_id=producto.id)
    )
    if almacen:
        q_salidas = q_salidas.filter(salida__almacen_id=almacen.id)

    if almacen:
        q_salidas = q_salidas.filter(salida__almacen_id=almacen.id)

    if dt_ini:
        q_salidas = q_salidas.filter(salida__fecha__gte=dt_ini)
    if dt_fin:
        q_salidas = q_salidas.filter(salida__fecha__lte=dt_fin)

    # Unificamos a una lista "movimientos"
    movs = []

    for d in q_entradas:        
        cantidad = d.cantidad or Decimal("0")
        unitario = d.costo_unitario
        importe = (cantidad * unitario) if unitario is not None else None
        almacen_real = d.almacen or d.entrada.almacen    


        movs.append({
            "fecha": d.entrada.fecha,
            "creado_en": d.entrada.creado_en,
            "es_entrada": True,
            "tipo": d.entrada.tipo,
            "tipo_display": d.entrada.get_tipo_display(),
            "folio": d.entrada.folio,
            "almacen": almacen_real,
            "almacen_display": str(almacen_real) if almacen_real else "",
            "tercero": d.entrada.proveedor,
            "documento_referencia": d.entrada.documento_referencia or "",
            "cantidad_entrada": cantidad,
            "cantidad_salida": None,
            "unitario": unitario,
            "importe": importe,
        })

    for d in q_salidas:
        # SalidaInventarioDetalle tiene cantidad + precio_unitario :contentReference[oaicite:3]{index=3}
        cantidad = d.cantidad or Decimal("0")
        unitario = d.precio_unitario
        importe = (cantidad * unitario) if unitario is not None else None

        movs.append({
            "fecha": d.salida.fecha,
            "creado_en": d.salida.creado_en,
            "es_entrada": False,
            "tipo": d.salida.tipo,
            "tipo_display": d.salida.get_tipo_display(),
            "folio": d.salida.folio,
            "almacen": d.salida.almacen,
            "almacen_display": str(d.salida.almacen) if d.salida.almacen else "",
            "tercero": d.salida.cliente or d.salida.proveedor,
            "documento_referencia": d.salida.documento_referencia,
            "cantidad_entrada": None,
            "cantidad_salida": cantidad,
            "unitario": unitario,
            "importe": importe,
        })

    # Orden: por fecha y luego por creado_en (para que el saldo sea consistente)
    movs.sort(key=lambda x: (x["fecha"], x["creado_en"]))

    # Totales y saldo
    tot_entradas = Decimal("0")
    tot_salidas = Decimal("0")
    saldo = Decimal("0")  # si luego quieres saldo inicial real, aquí lo ajustas

    saldo_inicial = saldo

    for m in movs:
        if m["es_entrada"]:
            tot_entradas += (m["cantidad_entrada"] or Decimal("0"))
            saldo += (m["cantidad_entrada"] or Decimal("0"))
        else:
            tot_salidas += (m["cantidad_salida"] or Decimal("0"))
            saldo -= (m["cantidad_salida"] or Decimal("0"))

        m["saldo"] = saldo

    saldo_final = saldo
    movimientos = movs

    return render(request, "inventarios/kardex.html", {
        "productos": productos,
        "almacenes": almacenes_qs,
        "almacen": almacen,
        "almacen_id": almacen_id,
        "producto": producto,
        "producto_id": producto.id,
        "fecha_ini": fecha_ini,
        "fecha_fin": fecha_fin,
        "movimientos": movimientos,
        "saldo_inicial": saldo_inicial,
        "saldo_final": saldo_final,
        "tot_entradas": tot_entradas,
        "tot_salidas": tot_salidas,
    })

@permiso_requerido("inventarios.view_inventariostock")
def kardex_export(request):
    producto_id = (request.GET.get("producto") or "").strip()
    almacenes_qs = Almacen.objects.filter(es_activo=True).order_by("tipo", "nombre")
    almacen = get_almacen_default()
    almacen_id = (request.GET.get("almacen") or "").strip()
    if almacen_id.isdigit():
        almacen = almacenes_qs.filter(pk=int(almacen_id)).first() or almacen
    fecha_ini = (request.GET.get("fecha_ini") or "").strip()
    fecha_fin = (request.GET.get("fecha_fin") or "").strip()

    if not producto_id:
        return HttpResponse("Falta parámetro producto.", status=400)

    producto = get_object_or_404(Producto, pk=producto_id)

    # Parse fechas
    dt_ini = None
    dt_fin = None
    try:
        if fecha_ini:
            dt_ini = datetime.strptime(fecha_ini, "%Y-%m-%d").date()
        if fecha_fin:
            dt_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    except ValueError:
        dt_ini = None
        dt_fin = None

    # ENTRADAS
    q_entradas = (
        EntradaInventarioDetalle.objects
        .select_related("entrada")
        .filter(producto_id=producto.id)
    )
    if almacen:
        q_entradas = q_entradas.filter(entrada__almacen_id=almacen.id)
    if dt_ini:
        q_entradas = q_entradas.filter(entrada__fecha__gte=dt_ini)
    if dt_fin:
        q_entradas = q_entradas.filter(entrada__fecha__lte=dt_fin)

    # SALIDAS
    q_salidas = (
        SalidaInventarioDetalle.objects
        .select_related("salida")
        .filter(producto_id=producto.id)
    )
    if almacen:
        q_salidas = q_salidas.filter(salida__almacen_id=almacen.id)
    if dt_ini:
        q_salidas = q_salidas.filter(salida__fecha__gte=dt_ini)
    if dt_fin:
        q_salidas = q_salidas.filter(salida__fecha__lte=dt_fin)

    movs = []

    for d in q_entradas:
        cantidad = d.cantidad or Decimal("0")
        unitario = d.costo_unitario
        importe = (cantidad * unitario) if unitario is not None else None
        almacen_real = d.almacen or d.entrada.almacen

        movs.append({
            "fecha": d.entrada.fecha,
            "creado_en": d.entrada.creado_en,
            "movimiento": "Entrada",
            "tipo": d.entrada.get_tipo_display(),
            "folio": d.entrada.folio,
            "almacen": almacen_real,
            "almacen_display": str(almacen_real) if almacen_real else "",
            "tercero": str(d.entrada.proveedor) if d.entrada.proveedor else "",
            "referencia": d.entrada.documento_referencia or "",
            "entrada": cantidad,
            "salida": None,
            "unitario": unitario,
            "importe": importe,
        })

    for d in q_salidas:
        cantidad = d.cantidad or Decimal("0")
        unitario = d.precio_unitario
        importe = (cantidad * unitario) if unitario is not None else None

        tercero = ""
        if d.salida.cliente:
            tercero = str(d.salida.cliente)
        elif d.salida.proveedor:
            tercero = str(d.salida.proveedor)

        movs.append({
            "fecha": d.salida.fecha,
            "creado_en": d.salida.creado_en,
            "movimiento": "Salida",
            "tipo": d.salida.get_tipo_display(),
            "folio": d.salida.folio,
            "almacen": d.salida.almacen,
            "almacen_display": str(d.salida.almacen) if d.salida.almacen else "",
            "tercero": tercero,
            "referencia": d.salida.documento_referencia or "",
            "entrada": None,
            "salida": cantidad,
            "unitario": unitario,
            "importe": importe,
        })

    movs.sort(key=lambda x: (x["fecha"], x["creado_en"]))

    # Calcular saldo y totales
    saldo = Decimal("0")
    tot_entradas = Decimal("0")
    tot_salidas = Decimal("0")

    for m in movs:
        if m["movimiento"] == "Entrada":
            tot_entradas += (m["entrada"] or Decimal("0"))
            saldo += (m["entrada"] or Decimal("0"))
        else:
            tot_salidas += (m["salida"] or Decimal("0"))
            saldo -= (m["salida"] or Decimal("0"))
        m["saldo"] = saldo

    saldo_final = saldo

    # ---------------- EXCEL ----------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Kardex"

    # Encabezado superior
    ws["A1"] = "KARDEX"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Producto:"
    ws["B2"] = producto.nombre
    ws["A3"] = "Rango:"
    ws["B3"] = f"{fecha_ini or '—'} a {fecha_fin or '—'}"

    ws["A4"] = "Entradas:"
    ws["B4"] = float(tot_entradas)
    ws["C4"] = "Salidas:"
    ws["D4"] = float(tot_salidas)
    ws["E4"] = "Saldo final:"
    ws["F4"] = float(saldo_final)

    for cell in ["A2","A3","A4","C4","E4"]:
        ws[cell].font = Font(bold=True)

    # Cabeceras de tabla
    headers = [
        "Fecha", "Movimiento", "Tipo", "Folio", "Tercero", "Referencia",
        "Entrada", "Salida", "Unitario", "Importe", "Saldo"
    ]
    start_row = 6
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Filas
    row = start_row + 1
    for m in movs:
        ws.cell(row=row, column=1, value=m["fecha"])
        ws.cell(row=row, column=2, value=m["movimiento"])
        ws.cell(row=row, column=3, value=m["tipo"])
        ws.cell(row=row, column=4, value=m["folio"])
        ws.cell(row=row, column=5, value=m["tercero"])
        ws.cell(row=row, column=6, value=m["referencia"])

        ws.cell(row=row, column=7, value=float(m["entrada"]) if m["entrada"] is not None else None)
        ws.cell(row=row, column=8, value=float(m["salida"]) if m["salida"] is not None else None)
        ws.cell(row=row, column=9, value=float(m["unitario"]) if m["unitario"] is not None else None)
        ws.cell(row=row, column=10, value=float(m["importe"]) if m["importe"] is not None else None)
        ws.cell(row=row, column=11, value=float(m["saldo"]) if m["saldo"] is not None else None)

        row += 1

    # Formatos
    # Fecha
    for r in range(start_row + 1, row):
        ws.cell(r, 1).number_format = "dd/mm/yyyy"
        # columnas numéricas
        for c in [7, 8, 11]:
            ws.cell(r, c).number_format = "#,##0.0000"
        for c in [9, 10]:
            ws.cell(r, c).number_format = "$#,##0.00"

    # Ajustar anchos
    widths = {
        1: 12, 2: 12, 3: 18, 4: 14, 5: 28, 6: 22,
        7: 12, 8: 12, 9: 12, 10: 14, 11: 12
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Exportar
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    safe_name = "".join(ch for ch in producto.nombre if ch.isalnum() or ch in (" ", "_", "-")).strip()
    safe_name = safe_name.replace(" ", "_") or f"producto_{producto.id}"
    filename = f"kardex_{safe_name}.xlsx"

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
