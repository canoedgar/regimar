# catalogos/views.py
from openpyxl import load_workbook
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from django.http import JsonResponse, HttpResponseForbidden

from django.core.exceptions import ValidationError

from django.shortcuts import render, redirect, get_object_or_404
from .models import Producto, Categoria, Proveedor, Proyecto, Cliente, Almacen, ParametroSistema, ClienteProductoPrecio
from .forms import ProductoForm, CategoriaForm, ProveedorForm, ProyectoForm, ClienteForm, AlmacenForm, ProductoMetricaConversionFormSet, ParametroSistemaForm, ClienteProductoPrecioForm
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test, login_required
from accounts.decorators import ADMIN_GROUP_NAME, administrador_requerido, grupos_requeridos, permiso_requerido

from django.db.models import Q
from django.db import IntegrityError
from django.views.decorators.http import require_POST

from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils import timezone

from catalogos.services.precios import (
    registrar_bitacora_precio_producto,
    registrar_historial_precio_producto,
)
from catalogos.services.constancia_situacion_fiscal import (
    ConstanciaFiscalPDFError,
    extraer_datos_constancia_situacion_fiscal,
)
from catalogos.services.regimenes_fiscales import codigos_regimenes_fiscales, regimen_fiscal_a_json

# Inicio Categorías

@permiso_requerido("catalogos.view_categoria")
def categorias_list(request):
    categorias = Categoria.objects.all().order_by("nombre")
    return render(request, "catalogos/categorias_list.html", {
        "categorias": categorias,
    })


@permiso_requerido("catalogos.add_categoria")
def categorias_create(request):
    if request.method == "POST":
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("categorias_list")
    else:
        form = CategoriaForm()

    return render(request, "catalogos/categorias_form.html", {
        "form": form,
        "modo_edicion": False,
    })


@permiso_requerido("catalogos.change_categoria")
def categorias_edit(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)

    if request.method == "POST":
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            return redirect("categorias_list")
    else:
        form = CategoriaForm(instance=categoria)

    return render(request, "catalogos/categorias_form.html", {
        "form": form,
        "modo_edicion": True,
        "categoria": categoria,
    })

# Fin Categorías

# Inicio Productos

def _parametro_decimal(clave, default="0"):
    parametro = ParametroSistema.objects.filter(clave=clave, activo=True).first()
    if not parametro:
        return Decimal(str(default))
    try:
        return Decimal(str(parametro.valor))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(str(default))


def _redondear_cajas(value):
    return Decimal(value or 0).quantize(Decimal("1"), rounding=ROUND_HALF_UP)


def _preparar_stock_cajas(producto, promedio_peso_variable):
    stock = Decimal(producto.stock or 0)
    factor = None

    if producto.maneja_peso_variable:
        factor = promedio_peso_variable
    else:
        conversiones = list(producto.conversiones_metricas.all())
        conversion = next((c for c in conversiones if c.activo), conversiones[0] if conversiones else None)
        if conversion:
            factor = conversion.factor_conversion

    if not factor or Decimal(factor) <= 0:
        producto.stock_cajas = None
        producto.stock_cajas_disponible = False
        return

    producto.stock_cajas = _redondear_cajas(stock / Decimal(factor))
    producto.stock_cajas_disponible = True


@permiso_requerido("catalogos.view_producto")
def productos_list(request):
    productos = Producto.objects.all().prefetch_related("conversiones_metricas").order_by("nombre")
    promedio_peso_variable = _parametro_decimal("ARRACHERA_PROM_CAJA", "0")

    for producto in productos:
        _preparar_stock_cajas(producto, promedio_peso_variable)

    return render(request, "catalogos/productos_list.html", {
        "productos": productos
    })


def _build_conversion_formset(request, producto=None):
    kwargs = {"instance": producto or Producto()}
    if request.method == "POST":
        kwargs.update({"data": request.POST, "files": request.FILES})
    formset = ProductoMetricaConversionFormSet(**kwargs)
    for form in formset.forms:
        form.producto = producto
    return formset


@permiso_requerido("catalogos.add_producto")
def productos_create(request):
    producto = Producto()

    if request.method == "POST":
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        formset = _build_conversion_formset(request, producto)

        if form.is_valid() and formset.is_valid():
            producto = form.save()
            formset.instance = producto
            conversiones = formset.save(commit=False)
            for conversion in conversiones:
                conversion.producto = producto
                conversion.save()
            for eliminado in formset.deleted_objects:
                if eliminado.pk:
                    eliminado.delete()
            registrar_bitacora_precio_producto(
                producto,
                usuario=request.user,
                motivo="Alta de producto",
            )
            messages.success(request, "Producto creado correctamente.")
            return redirect("productos_list")
    else:
        form = ProductoForm(instance=producto)
        formset = _build_conversion_formset(request, producto)

    return render(request, "catalogos/productos_form.html", {
        "form": form,
        "formset_conversiones": formset,
        "producto": producto,
    })



@permiso_requerido("catalogos.change_producto")
def productos_edit(request, pk):
    producto = get_object_or_404(Producto, pk=pk)

    if request.method == "POST":
        precio_anterior = producto.precio
        precio_minimo_anterior = producto.precio_minimo

        form = ProductoForm(request.POST, request.FILES, instance=producto)
        formset = _build_conversion_formset(request, producto)
        if form.is_valid() and formset.is_valid():
            producto = form.save(commit=False)

            if producto.precio != precio_anterior:
                producto.fecha_ultima_actualizacion_precio = timezone.now()

            producto.save()

            registrar_historial_precio_producto(
                producto=producto,
                precio_anterior=precio_anterior,
                precio_nuevo=producto.precio,
                precio_minimo_anterior=precio_minimo_anterior,
                precio_minimo_nuevo=producto.precio_minimo,
                usuario=request.user,
                motivo="Actualización desde catálogo de productos",
            )
            registrar_bitacora_precio_producto(
                producto,
                usuario=request.user,
                motivo="Actualización desde catálogo de productos",
            )

            formset.instance = producto
            conversiones = formset.save(commit=False)
            for conversion in conversiones:
                conversion.producto = producto
                conversion.save()
            for eliminado in formset.deleted_objects:
                if eliminado.pk:
                    eliminado.delete()
            messages.success(request, "Producto actualizado correctamente.")
            return redirect("productos_list")
    else:
        form = ProductoForm(instance=producto)
        formset = _build_conversion_formset(request, producto)

    return render(request, "catalogos/productos_form.html", {
        "form": form,
        "formset_conversiones": formset,
        "producto": producto,
    })

@permiso_requerido("catalogos.delete_producto")
def productos_delete(request, pk):
    producto = get_object_or_404(Producto, pk=pk)

    if producto.stock != 0:
        messages.error(request, f"No se puede eliminar un producto '{producto.nombre}' porque su stock es {producto.stock}.")
        return redirect("productos_list")

    if request.method == "POST":
        nombre = producto.nombre
        producto.delete()        
        messages.success(request, f"Producto '{nombre}' eliminado correctamente.")
        return redirect("productos_list")

    return render(request, "catalogos/productos_confirm_delete.html", {
        "producto": producto,        
    })




@permiso_requerido("catalogos.view_producto", "catalogos.change_producto")
def precios_productos_list(request):
    productos = Producto.objects.all().order_by("nombre")

    if request.method == "POST":
        producto_id = request.POST.get("producto_id")
        producto = get_object_or_404(Producto, pk=producto_id)

        precio_anterior = producto.precio
        precio_minimo_anterior = producto.precio_minimo

        try:
            nuevo_precio = Decimal(str(request.POST.get("precio", "0") or "0"))
            nuevo_precio_minimo = Decimal(str(request.POST.get("precio_minimo", "0") or "0"))
        except (InvalidOperation, ValueError, TypeError):
            messages.error(request, "Precio inválido.")
            return redirect("precios_productos_list")

        if nuevo_precio < 0 or nuevo_precio_minimo < 0:
            messages.error(request, "Los precios no pueden ser negativos.")
            return redirect("precios_productos_list")

        producto.precio = nuevo_precio
        producto.precio_minimo = nuevo_precio_minimo
        producto.fecha_ultima_actualizacion_precio = timezone.now()
        producto.save(update_fields=["precio", "precio_minimo", "fecha_ultima_actualizacion_precio"])

        registrar_historial_precio_producto(
            producto=producto,
            precio_anterior=precio_anterior,
            precio_nuevo=producto.precio,
            precio_minimo_anterior=precio_minimo_anterior,
            precio_minimo_nuevo=producto.precio_minimo,
            usuario=request.user,
            motivo="Actualización desde lista maestra de precios",
        )
        registrar_bitacora_precio_producto(
            producto,
            usuario=request.user,
            motivo="Actualización desde lista maestra de precios",
        )

        messages.success(request, f"Precio de {producto.nombre} actualizado correctamente.")
        return redirect("precios_productos_list")

    return render(request, "catalogos/precios_productos_list.html", {
        "productos": productos,
    })


@permiso_requerido("catalogos.view_producto")
def producto_precio_bitacora(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    bitacora = producto.bitacora_precios.all()[:60]
    historial = producto.historial_precios.all()[:60]

    return render(request, "catalogos/producto_precio_bitacora.html", {
        "producto": producto,
        "bitacora": bitacora,
        "historial": historial,
    })


# --- Parámetros de sistema ---

def _es_admin(user):
    return user.is_authenticated and (user.is_superuser or user.groups.filter(name="Administrador").exists())


def _asegurar_parametros_base():
    ParametroSistema.objects.get_or_create(
        clave="PRECIO_VIGENCIA_DIAS",
        defaults={
            "nombre": "Días máximos de vigencia del último precio por cliente",
            "valor": "30",
            "descripcion": "Regla general para advertir cuando un cliente lleva demasiados días sin comprar un producto al último precio otorgado.",
            "activo": True,
        },
    )


@permiso_requerido("catalogos.view_parametrosistema")
def parametros_sistema_list(request):
    _asegurar_parametros_base()
    parametros = ParametroSistema.objects.all().order_by("clave")
    return render(request, "catalogos/parametros_sistema_list.html", {"parametros": parametros})


@permiso_requerido("catalogos.add_parametrosistema")
def parametros_sistema_create(request):
    form = ParametroSistemaForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Parámetro creado correctamente.")
        return redirect("parametros_sistema_list")
    return render(request, "catalogos/parametros_sistema_form.html", {"form": form, "modo": "crear"})


@permiso_requerido("catalogos.change_parametrosistema")
def parametros_sistema_edit(request, pk):
    parametro = get_object_or_404(ParametroSistema, pk=pk)
    form = ParametroSistemaForm(request.POST or None, instance=parametro)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Parámetro actualizado correctamente.")
        return redirect("parametros_sistema_list")
    return render(request, "catalogos/parametros_sistema_form.html", {"form": form, "parametro": parametro, "modo": "editar"})


@permiso_requerido("catalogos.view_clienteproductoprecio")
def precios_clientes_list(request):
    precios = (
        ClienteProductoPrecio.objects
        .select_related("cliente", "producto", "actualizado_por")
        .order_by("cliente__nombre_fiscal", "producto__nombre")
    )
    q = (request.GET.get("q") or "").strip()
    if q:
        precios = precios.filter(
            Q(cliente__nombre_fiscal__icontains=q) |
            Q(cliente__nombre_comercial__icontains=q) |
            Q(producto__nombre__icontains=q)
        )
    return render(request, "catalogos/precios_clientes_list.html", {"precios": precios, "q": q})


@permiso_requerido("catalogos.change_clienteproductoprecio")
def precio_cliente_edit(request, pk):
    precio_cliente = get_object_or_404(ClienteProductoPrecio.objects.select_related("cliente", "producto"), pk=pk)
    precio_anterior = precio_cliente.ultimo_precio
    form = ClienteProductoPrecioForm(request.POST or None, instance=precio_cliente)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.precio_anterior = precio_anterior
        obj.actualizado_por = request.user
        obj.fecha_ultimo_precio = timezone.now()
        obj.save()
        messages.success(request, "Último precio del cliente actualizado correctamente.")
        return redirect("precios_clientes_list")
    return render(request, "catalogos/precio_cliente_form.html", {"form": form, "precio_cliente": precio_cliente})

# Fin Productos


# Inicio Proveedores

@permiso_requerido("catalogos.view_proveedor")
def proveedores_list(request):
    proveedores = Proveedor.objects.all().order_by("nombre")
    return render(request, "catalogos/proveedores_list.html", {
        "proveedores": proveedores
    })

@permiso_requerido("catalogos.add_proveedor")
def proveedores_create(request):
    next_url = (request.GET.get("next") or "").strip()

    if request.method == "POST":
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save()
            messages.success(request, "Proveedor creado correctamente.")

            # ✅ Si venías de otra pantalla (ej. entrada manual), regresa y selecciona el proveedor
            if next_url:
                sep = "&" if "?" in next_url else "?"
                return redirect(f"{next_url}{sep}{urlencode({'proveedor_id': proveedor.id})}")

            return redirect("proveedores_list")
    else:
        form = ProveedorForm()

    return render(request, "catalogos/proveedores_form.html", {
        "form": form,
        "next": next_url,  # para que el template pueda usarlo en Cancelar
    })

@permiso_requerido("catalogos.change_proveedor")
def proveedores_edit(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == "POST":
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            return redirect("proveedores_list")
    else:
        form = ProveedorForm(instance=proveedor)

    return render(request, "catalogos/proveedores_form.html", {
        "form": form,
        "proveedor": proveedor,
    })

# Fin Proveedores

# --- Proyectos ---

@permiso_requerido("catalogos.view_proyecto")
def proyectos_list(request):
    estado = request.GET.get("estado", "").strip()
    qs = Proyecto.objects.all().order_by("-fecha_actualizacion", "nombre")
    if estado:
        qs = qs.filter(estado=estado)

    return render(request, "catalogos/proyectos_list.html", {
        "proyectos": qs,
        "estado_actual": estado,
        "ESTADOS": Proyecto.Estado.choices,
    })


@permiso_requerido("catalogos.add_proyecto")
def proyectos_create(request):
    if request.method == "POST":
        form = ProyectoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Proyecto creado correctamente.")
            return redirect("proyectos_list")
    else:
        form = ProyectoForm()

    return render(request, "catalogos/proyectos_form.html", {
        "form": form,
        "modo_edicion": False,
    })


@permiso_requerido("catalogos.change_proyecto")
def proyectos_edit(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)

    if request.method == "POST":
        form = ProyectoForm(request.POST, instance=proyecto)
        if form.is_valid():
            form.save()
            messages.success(request, "Proyecto actualizado correctamente.")
            return redirect("proyectos_list")
    else:
        form = ProyectoForm(instance=proyecto)

    return render(request, "catalogos/proyectos_form.html", {
        "form": form,
        "modo_edicion": True,
        "proyecto": proyecto,
    })

# --- Fin Proyectos ---

# --- Inicio Clientes ---

def _puede_editar_parametros_cartera(user):
    return bool(
        user
        and user.is_authenticated
        and (user.is_superuser or user.groups.filter(name=ADMIN_GROUP_NAME).exists())
    )


def _get_safe_next_url(request, default_url_name="clientes_list"):
    next_url = request.POST.get("next") or request.GET.get("next") or ""

    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure()
    ):
        return next_url

    return reverse(default_url_name)


def _add_cliente_to_next_url(next_url, cliente_id):
    """Devuelve al flujo origen conservando querystring y marcando el cliente recién creado."""
    parts = urlsplit(next_url)
    query_params = dict(parse_qsl(parts.query, keep_blank_values=True))
    query_params["cliente_id"] = str(cliente_id)
    return urlunsplit((
        parts.scheme,
        parts.netloc,
        parts.path,
        urlencode(query_params),
        parts.fragment,
    ))


def _aplicar_datos_constancia_en_post(post_data, datos_constancia):
    """Mezcla datos extraídos de CSF en un POST mutable para re-renderizar el form."""
    data = post_data.copy()
    campos_cliente = [
        "tipo_persona",
        "rfc",
        "nombre_fiscal",
        "nombre_comercial",
        "domicilio_fiscal_cp",
        "calle",
        "num_ext",
        "num_int",
        "colonia",
        "localidad",
        "municipio",
        "estado",
        "pais",
        "cp",
        "referencias",
    ]

    regimenes = datos_constancia.get("regimenes_fiscales_codigos") or codigos_regimenes_fiscales(
        datos_constancia.get("regimen_fiscal")
    )
    if regimenes:
        data.setlist("regimen_fiscal", regimenes)

    for campo in campos_cliente:
        valor = (datos_constancia.get(campo) or "").strip()
        if valor:
            data[campo] = valor

    return data


def _leer_constancia_desde_request(request):
    archivo = request.FILES.get("constancia_situacion_fiscal_pdf")
    if not archivo:
        raise ConstanciaFiscalPDFError("Selecciona un PDF de constancia de situación fiscal para leerlo.")
    return extraer_datos_constancia_situacion_fiscal(archivo)


def _debe_abrir_datos_fiscales(form=None, forzar=False):
    if forzar:
        return True
    if not form or not getattr(form, "errors", None):
        return False

    campos_fiscales = {
        "tipo_persona",
        "rfc",
        "nombre_fiscal",
        "regimen_fiscal",
        "domicilio_fiscal_cp",
        "uso_cfdi_default",
        "email_cfdi",
        "constancia_situacion_fiscal_pdf",
    }
    return bool(campos_fiscales.intersection(form.errors.keys()) or form.non_field_errors())


def _mensaje_constancia_leida(request, datos_constancia):
    campos_importantes = [
        "rfc",
        "nombre_fiscal",
        "regimenes_fiscales_codigos",
        "domicilio_fiscal_cp",
        "calle",
        "colonia",
        "municipio",
        "estado",
    ]
    total = sum(1 for campo in campos_importantes if datos_constancia.get(campo))
    total_regimenes = len(datos_constancia.get("regimenes_fiscales_codigos") or [])
    detalle_regimenes = f" Se detectaron {total_regimenes} régimen(es) fiscal(es)." if total_regimenes else ""
    messages.success(
        request,
        f"Constancia leída correctamente. Se precargaron {total} campos; revisa la información antes de guardar.{detalle_regimenes}",
    )


@permiso_requerido("catalogos.view_cliente")
def clientes_list(request):
    q = (request.GET.get("q") or "").strip()

    clientes = Cliente.objects.all()
    if q:
        clientes = clientes.filter(
            Q(rfc__icontains=q) |
            Q(nombre_fiscal__icontains=q) |
            Q(nombre_comercial__icontains=q)
        )

    context = {
        "clientes": clientes,
        "q": q,
        "total": clientes.count(),
    }
    return render(request, "catalogos/clientes_list.html", context)


@permiso_requerido("catalogos.add_cliente")
def cliente_create(request):
    next_url = _get_safe_next_url(request)
    puede_editar_parametros_cartera = _puede_editar_parametros_cartera(request.user)
    abrir_datos_fiscales = False

    if request.method == "POST":
        accion = request.POST.get("accion") or "guardar"

        if accion == "extraer_constancia":
            data_form = request.POST
            abrir_datos_fiscales = True
            try:
                datos_constancia = _leer_constancia_desde_request(request)
                data_form = _aplicar_datos_constancia_en_post(request.POST, datos_constancia)
                _mensaje_constancia_leida(request, datos_constancia)
            except ConstanciaFiscalPDFError as exc:
                messages.error(request, str(exc))

            form = ClienteForm(
                data_form,
                request.FILES,
                puede_editar_parametros_cartera=puede_editar_parametros_cartera,
            )
        else:
            form = ClienteForm(
                request.POST,
                request.FILES,
                puede_editar_parametros_cartera=puede_editar_parametros_cartera,
            )
            if form.is_valid():
                try:
                    cliente = form.save()
                    messages.success(request, "Cliente creado correctamente.")
                    return redirect(_add_cliente_to_next_url(next_url, cliente.id))
                except IntegrityError:
                    form.add_error("rfc", "Ya existe un cliente con ese RFC.")
            abrir_datos_fiscales = _debe_abrir_datos_fiscales(form)
    else:
        form = ClienteForm(puede_editar_parametros_cartera=puede_editar_parametros_cartera)

    return render(request, "catalogos/cliente_form.html", {
        "form": form,
        "modo": "crear",
        "next_url": next_url,
        "puede_editar_parametros_cartera": puede_editar_parametros_cartera,
        "abrir_datos_fiscales": abrir_datos_fiscales,
    })

@permiso_requerido("catalogos.change_cliente")
def cliente_edit(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    puede_editar_parametros_cartera = _puede_editar_parametros_cartera(request.user)
    abrir_datos_fiscales = False

    if request.method == "POST":
        accion = request.POST.get("accion") or "guardar"

        if accion == "extraer_constancia":
            data_form = request.POST
            abrir_datos_fiscales = True
            try:
                datos_constancia = _leer_constancia_desde_request(request)
                data_form = _aplicar_datos_constancia_en_post(request.POST, datos_constancia)
                _mensaje_constancia_leida(request, datos_constancia)
            except ConstanciaFiscalPDFError as exc:
                messages.error(request, str(exc))

            form = ClienteForm(
                data_form,
                request.FILES,
                instance=cliente,
                puede_editar_parametros_cartera=puede_editar_parametros_cartera,
            )
        else:
            form = ClienteForm(
                request.POST,
                request.FILES,
                instance=cliente,
                puede_editar_parametros_cartera=puede_editar_parametros_cartera,
            )
            if form.is_valid():
                try:
                    form.save()
                    messages.success(request, "Cliente actualizado correctamente.")
                    return redirect("clientes_list")
                except IntegrityError:
                    form.add_error("rfc", "Ya existe un cliente con ese RFC.")
            abrir_datos_fiscales = _debe_abrir_datos_fiscales(form)
    else:
        form = ClienteForm(
            instance=cliente,
            puede_editar_parametros_cartera=puede_editar_parametros_cartera,
        )

    return render(request, "catalogos/cliente_form.html", {
        "form": form,
        "modo": "editar",
        "cliente": cliente,
        "puede_editar_parametros_cartera": puede_editar_parametros_cartera,
        "abrir_datos_fiscales": abrir_datos_fiscales,
    })


@permiso_requerido("catalogos.add_cliente")
@require_POST
def cliente_quick_create(request):
    try:
        rfc = (request.POST.get("rfc") or "").strip().upper()
        nombre_fiscal = (request.POST.get("nombre_fiscal") or "").strip()
        regimen_fiscal = (request.POST.get("regimen_fiscal") or "").strip()
        domicilio_fiscal_cp = (request.POST.get("domicilio_fiscal_cp") or "").strip()

        telefono = (request.POST.get("telefono") or "").strip()
        contacto = (request.POST.get("contacto") or "").strip()
        email_cfdi = (request.POST.get("email_cfdi") or "").strip()

        if not (rfc and nombre_fiscal and regimen_fiscal and domicilio_fiscal_cp):
            return JsonResponse({"ok": False, "error": "Faltan campos requeridos."}, status=400)

        c = Cliente(
            rfc=rfc,
            nombre_fiscal=nombre_fiscal,
            regimen_fiscal=regimen_fiscal_a_json([regimen_fiscal]),
            domicilio_fiscal_cp=domicilio_fiscal_cp,
            telefono=telefono,
            contacto=contacto,
            email_cfdi=email_cfdi,
        )

        c.full_clean()
        c.save()

        return JsonResponse({
            "ok": True,
            "id": c.id,
            "rfc": c.rfc,
            "nombre_fiscal": c.nombre_fiscal,
            "telefono": c.telefono,
            "contacto": c.contacto,
            "email_cfdi": c.email_cfdi,
        })

    except ValidationError as e:
        return JsonResponse({"ok": False, "error": "; ".join(sum(e.message_dict.values(), []))}, status=400)
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)


# --- Fin Clientes ---

# --- Almacenes ---

@permiso_requerido("catalogos.view_almacen")
def almacenes_list(request):
    q = (request.GET.get("q") or "").strip()

    almacenes = Almacen.objects.all().order_by("nombre")
    if q:
        almacenes = almacenes.filter(
            Q(codigo__icontains=q) |
            Q(nombre__icontains=q) |
            Q(tipo__icontains=q)
        )

    return render(request, "catalogos/almacenes_list.html", {
        "almacenes": almacenes,
        "q": q,
        "total": almacenes.count(),
    })


@permiso_requerido("catalogos.add_almacen")
def almacenes_create(request):
    if request.method == "POST":
        form = AlmacenForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Almacén creado correctamente.")
            return redirect("almacenes_list")
    else:
        form = AlmacenForm()

    return render(request, "catalogos/almacenes_form.html", {
        "form": form,
        "modo_edicion": False,
    })


@permiso_requerido("catalogos.change_almacen")
def almacenes_edit(request, pk):
    almacen = get_object_or_404(Almacen, pk=pk)

    if request.method == "POST":
        form = AlmacenForm(request.POST, instance=almacen)
        if form.is_valid():
            form.save()
            messages.success(request, "Almacén actualizado correctamente.")
            return redirect("almacenes_list")
    else:
        form = AlmacenForm(instance=almacen)

    return render(request, "catalogos/almacenes_form.html", {
        "form": form,
        "modo_edicion": True,
        "almacen": almacen,
    })


@permiso_requerido("catalogos.delete_almacen")
def almacenes_confirm_delete(request, pk):
    almacen = get_object_or_404(Almacen, pk=pk)

    if request.method == "POST":
        nombre = str(almacen)
        almacen.delete()
        messages.success(request, f"Almacén '{nombre}' eliminado correctamente.")
        return redirect("almacenes_list")

    return render(request, "catalogos/almacenes_confirm_delete.html", {
        "almacen": almacen,
    })

# --- Fin Almacenes ---

# --- Inicio Importaciones ---

def _to_str(v):
    if v is None:
        return ""
    return str(v).strip()

def _to_bool_si_no(v, default=True):
    s = _to_str(v).upper()
    if s in ("SI", "S", "TRUE", "1", "X"):
        return True
    if s in ("NO", "N", "FALSE", "0"):
        return False
    return default

def _to_decimal(v, default=Decimal("0")):
    if v is None or str(v).strip() == "":
        return default
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        return default

def _sheet_as_dict_rows(ws):
    """
    Lee la hoja usando encabezados en la fila 1.
    Devuelve lista de dicts: {header: value}.
    """
    headers = [(_to_str(c.value)) for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
        rows.append(row)
    return headers, rows

@permiso_requerido("catalogos.add_producto", "catalogos.change_producto")
def importar_productos(request):
    if request.method == "POST" and request.FILES.get("archivo"):
        creados, actualizados, saltados = 0, 0, 0
        errores = []

        try:
            wb = load_workbook(request.FILES["archivo"], data_only=True)
            ws = wb.active
            headers, rows = _sheet_as_dict_rows(ws)

            # Esperados según tu template base
            # categoria, nombre, clave_sat, metrica, precio, stock, stock_minimo, stock_maximo, imagen, imagen_base64

            for idx, row in enumerate(rows, start=2):
                categoria_nombre = _to_str(row.get("categoria"))
                nombre = _to_str(row.get("nombre"))
                if not nombre:
                    saltados += 1
                    continue

                # categoria puede venir vacío porque en tu modelo es null=True/blank=True【turn2file2†models.py†L47-L54】
                categoria = None
                if categoria_nombre:
                    categoria, _ = Categoria.objects.get_or_create(nombre=categoria_nombre)

                data = {
                    "categoria": categoria,
                    "nombre": nombre,
                    "clave_sat": _to_str(row.get("clave_sat")) or None,
                    "metrica": _to_str(row.get("metrica")) or "PIEZA",
                    "precio": _to_decimal(row.get("precio")),
                    "stock": _to_decimal(row.get("stock")),
                    "stock_minimo": _to_decimal(row.get("stock_minimo")),
                    "stock_maximo": _to_decimal(row.get("stock_maximo")),
                    "imagen_base64": _to_str(row.get("imagen_base64")) or None,
                }

                # Nota: "imagen" (ImageField) no conviene importarla desde Excel directo (requiere archivo real).
                # Si quieres, luego hacemos una importación de imágenes por ZIP/rutas.

                try:
                    # En tu modelo, nombre_normalizado es unique, así que mejor update_or_create por nombre
                    obj, created = Producto.objects.update_or_create(
                        nombre=nombre,
                        defaults=data,
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {e}")

            if errores:
                messages.warning(
                    request,
                    f"Importación terminada. Creados: {creados}, Actualizados: {actualizados}, Saltados: {saltados}. "
                    f"Errores: {len(errores)} (ver consola)."
                )
                for err in errores[:20]:
                    print(err)
            else:
                messages.success(
                    request,
                    f"Importación terminada. Creados: {creados}, Actualizados: {actualizados}, Saltados: {saltados}."
                )

        except Exception as e:
            messages.error(request, f"Error al importar: {e}")

        return redirect("importar_productos")

    return render(request, "catalogos/importar_productos.html")

@permiso_requerido("catalogos.add_proveedor", "catalogos.change_proveedor")
def importar_proveedores(request):
    if request.method == "POST" and request.FILES.get("archivo"):
        creados, actualizados, saltados = 0, 0, 0
        errores = []

        try:
            wb = load_workbook(request.FILES["archivo"], data_only=True)
            ws = wb.active
            headers, rows = _sheet_as_dict_rows(ws)

            for idx, row in enumerate(rows, start=2):
                nombre = _to_str(row.get("nombre"))
                if not nombre:
                    saltados += 1
                    continue

                data = {
                    "nombre": nombre,
                    "rfc": _to_str(row.get("rfc")) or None,
                    "contacto": _to_str(row.get("contacto")) or None,
                    "telefono": _to_str(row.get("telefono")) or None,
                    "email": _to_str(row.get("email")) or None,
                    "direccion": _to_str(row.get("direccion")) or None,
                    "activo": _to_bool_si_no(row.get("activo"), default=True),
                }

                try:
                    # nombre_normalizado es unique => update_or_create por nombre
                    obj, created = Proveedor.objects.update_or_create(
                        nombre=nombre,
                        defaults=data,
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {e}")

            if errores:
                messages.warning(
                    request,
                    f"Importación terminada. Creados: {creados}, Actualizados: {actualizados}, Saltados: {saltados}. "
                    f"Errores: {len(errores)} (ver consola)."
                )
                for err in errores[:20]:
                    print(err)
            else:
                messages.success(
                    request,
                    f"Importación terminada. Creados: {creados}, Actualizados: {actualizados}, Saltados: {saltados}."
                )

        except Exception as e:
            messages.error(request, f"Error al importar: {e}")

        return redirect("importar_proveedores")

    return render(request, "catalogos/importar_proveedores.html")


@permiso_requerido("catalogos.add_cliente", "catalogos.change_cliente")
def importar_clientes(request):
    if request.method == "POST" and request.FILES.get("archivo"):
        creados, actualizados, saltados = 0, 0, 0
        errores = []

        try:
            wb = load_workbook(request.FILES["archivo"], data_only=True)
            ws = wb.active
            headers, rows = _sheet_as_dict_rows(ws)

            for idx, row in enumerate(rows, start=2):
                # Campos clave / obligatorios
                rfc = _to_str(row.get("rfc")).upper()
                nombre_fiscal = _to_str(row.get("nombre_fiscal"))
                regimen_fiscal = _to_str(row.get("regimen_fiscal"))
                domicilio_fiscal_cp = _to_str(row.get("domicilio_fiscal_cp"))

                # si falta algo obligatorio, se salta y lo reporta
                if not rfc or not nombre_fiscal or not regimen_fiscal or not domicilio_fiscal_cp:
                    saltados += 1
                    continue

                data = {
                    "tipo_persona": _to_str(row.get("tipo_persona")) or Cliente.TipoPersona.MORAL,
                    "rfc": rfc,
                    "nombre_fiscal": nombre_fiscal,
                    "regimen_fiscal": regimen_fiscal_a_json([regimen_fiscal]),
                    "domicilio_fiscal_cp": domicilio_fiscal_cp,

                    "uso_cfdi_default": _to_str(row.get("uso_cfdi_default")),
                    "email_cfdi": _to_str(row.get("email_cfdi")),
                    "nombre_comercial": _to_str(row.get("nombre_comercial")),
                    "telefono": _to_str(row.get("telefono")),

                    "calle": _to_str(row.get("calle")),
                    "num_ext": _to_str(row.get("num_ext")),
                    "num_int": _to_str(row.get("num_int")),
                    "colonia": _to_str(row.get("colonia")),
                    "localidad": _to_str(row.get("localidad")),
                    "municipio": _to_str(row.get("municipio")),
                    "estado": _to_str(row.get("estado")),
                    "pais": _to_str(row.get("pais")) or "México",
                    "cp": _to_str(row.get("cp")),
                    "referencias": _to_str(row.get("referencias")),

                    "forma_pago_default": _to_str(row.get("forma_pago_default")),
                    "metodo_pago_default": _to_str(row.get("metodo_pago_default")),
                    "activo": _to_bool_si_no(row.get("activo"), default=True),
                }

                try:
                    obj, created = Cliente.objects.update_or_create(
                        rfc=rfc,  # es unique【turn2file0†models.py†L34-L44】
                        defaults=data,
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                except Exception as e:
                    errores.append(f"Fila {idx} (RFC {rfc}): {e}")

            if errores:
                messages.warning(
                    request,
                    f"Importación terminada. Creados: {creados}, Actualizados: {actualizados}, Saltados: {saltados}. "
                    f"Errores: {len(errores)} (ver consola)."
                )
                for err in errores[:30]:
                    print(err)
            else:
                messages.success(
                    request,
                    f"Importación terminada. Creados: {creados}, Actualizados: {actualizados}, Saltados: {saltados}."
                )

        except Exception as e:
            messages.error(request, f"Error al importar: {e}")

        return redirect("importar_clientes")

    return render(request, "catalogos/importar_clientes.html")

# --- Fin Importaciones ---
