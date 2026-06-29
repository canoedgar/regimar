# catalogos/views.py
from openpyxl import load_workbook
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from django.http import JsonResponse, HttpResponseForbidden

from django.core.exceptions import ValidationError

from django.shortcuts import render, redirect, get_object_or_404
from .models import Producto, Categoria, Proveedor, Proyecto, Cliente, Almacen, ParametroSistema, ClienteProductoPrecio
from .forms import ProductoForm, CategoriaForm, ProveedorForm, ProyectoForm, ClienteForm, AlmacenForm, ProductoMetricaConversionFormSet, ParametroSistemaForm, ClienteProductoPrecioForm
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test, login_required
from accounts.decorators import ADMIN_GROUP_NAME, administrador_requerido, grupos_requeridos, permiso_requerido

from django.db.models import Q
from django.db import IntegrityError
from django.views.decorators.http import require_POST

from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils import timezone

from catalogos.services.precios import (
    registrar_bitacora_precio_producto,
    registrar_historial_precio_producto,
)
from catalogos.services.constancia_situacion_fiscal import (
    ConstanciaFiscalPDFError,
    extraer_datos_constancia_situacion_fiscal,
)
from catalogos.services.regimenes_fiscales import codigos_regimenes_fiscales, regimen_fiscal_a_json

# Inicio Categorías

def _to_str(v):
    if v is None:
        return ""
    return str(v).strip()

def _to_bool_si_no(v, default=True):
    s = _to_str(v).upper()
    if s in ("SI", "S", "TRUE", "1", "X"):
        return True
    if s in ("NO", "N", "FALSE", "0"):
        return False
    return default

def _to_decimal(v, default=Decimal("0")):
    if v is None or str(v).strip() == "":
        return default
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        return default

def _sheet_as_dict_rows(ws):
    """
    Lee la hoja usando encabezados en la fila 1.
    Devuelve lista de dicts: {header: value}.
    """
    headers = [(_to_str(c.value)) for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
        rows.append(row)
    return headers, rows

@permiso_requerido("catalogos.add_producto", "catalogos.change_producto")
def importar_productos(request):
    if request.method == "POST" and request.FILES.get("archivo"):
        creados, actualizados, saltados = 0, 0, 0
        errores = []

        try:
            wb = load_workbook(request.FILES["archivo"], data_only=True)
            ws = wb.active
            headers, rows = _sheet_as_dict_rows(ws)

            # Esperados según tu template base
            # categoria, nombre, clave_sat, metrica, precio, stock, stock_minimo, stock_maximo, imagen, imagen_base64

            for idx, row in enumerate(rows, start=2):
                categoria_nombre = _to_str(row.get("categoria"))
                nombre = _to_str(row.get("nombre"))
                if not nombre:
                    saltados += 1
                    continue

                # categoria puede venir vacío porque en tu modelo es null=True/blank=True【turn2file2†models.py†L47-L54】
                categoria = None
                if categoria_nombre:
                    categoria, _ = Categoria.objects.get_or_create(nombre=categoria_nombre)

                data = {
                    "categoria": categoria,
                    "nombre": nombre,
                    "clave_sat": _to_str(row.get("clave_sat")) or None,
                    "metrica": _to_str(row.get("metrica")) or "PIEZA",
                    "precio": _to_decimal(row.get("precio")),
                    "stock": _to_decimal(row.get("stock")),
                    "stock_minimo": _to_decimal(row.get("stock_minimo")),
                    "stock_maximo": _to_decimal(row.get("stock_maximo")),
                    "imagen_base64": _to_str(row.get("imagen_base64")) or None,
                }

                # Nota: "imagen" (ImageField) no conviene importarla desde Excel directo (requiere archivo real).
                # Si quieres, luego hacemos una importación de imágenes por ZIP/rutas.

                try:
                    # En tu modelo, nombre_normalizado es unique, así que mejor update_or_create por nombre
                    obj, created = Producto.objects.update_or_create(
                        nombre=nombre,
                        defaults=data,
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {e}")

            if errores:
                messages.warning(
                    request,
                    f"Importación terminada. Creados: {creados}, Actualizados: {actualizados}, Saltados: {saltados}. "
                    f"Errores: {len(errores)} (ver consola)."
                )
                for err in errores[:20]:
                    print(err)
            else:
                messages.success(
                    request,
                    f"Importación terminada. Creados: {creados}, Actualizados: {actualizados}, Saltados: {saltados}."
                )

        except Exception as e:
            messages.error(request, f"Error al importar: {e}")

        return redirect("importar_productos")

    return render(request, "catalogos/importar_productos.html")

@permiso_requerido("catalogos.add_proveedor", "catalogos.change_proveedor")
def importar_proveedores(request):
    if request.method == "POST" and request.FILES.get("archivo"):
        creados, actualizados, saltados = 0, 0, 0
        errores = []

        try:
            wb = load_workbook(request.FILES["archivo"], data_only=True)
            ws = wb.active
            headers, rows = _sheet_as_dict_rows(ws)

            for idx, row in enumerate(rows, start=2):
                nombre = _to_str(row.get("nombre"))
                if not nombre:
                    saltados += 1
                    continue

                data = {
                    "nombre": nombre,
                    "rfc": _to_str(row.get("rfc")) or None,
                    "contacto": _to_str(row.get("contacto")) or None,
                    "telefono": _to_str(row.get("telefono")) or None,
                    "email": _to_str(row.get("email")) or None,
                    "direccion": _to_str(row.get("direccion")) or None,
                    "activo": _to_bool_si_no(row.get("activo"), default=True),
                }

                try:
                    # nombre_normalizado es unique => update_or_create por nombre
                    obj, created = Proveedor.objects.update_or_create(
                        nombre=nombre,
                        defaults=data,
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {e}")

            if errores:
                messages.warning(
                    request,
                    f"Importación terminada. Creados: {creados}, Actualizados: {actualizados}, Saltados: {saltados}. "
                    f"Errores: {len(errores)} (ver consola)."
                )
                for err in errores[:20]:
                    print(err)
            else:
                messages.success(
                    request,
                    f"Importación terminada. Creados: {creados}, Actualizados: {actualizados}, Saltados: {saltados}."
                )

        except Exception as e:
            messages.error(request, f"Error al importar: {e}")

        return redirect("importar_proveedores")

    return render(request, "catalogos/importar_proveedores.html")


@permiso_requerido("catalogos.add_cliente", "catalogos.change_cliente")
def importar_clientes(request):
    if request.method == "POST" and request.FILES.get("archivo"):
        creados, actualizados, saltados = 0, 0, 0
        errores = []

        try:
            wb = load_workbook(request.FILES["archivo"], data_only=True)
            ws = wb.active
            headers, rows = _sheet_as_dict_rows(ws)

            for idx, row in enumerate(rows, start=2):
                # Campos clave / obligatorios
                rfc = _to_str(row.get("rfc")).upper()
                nombre_fiscal = _to_str(row.get("nombre_fiscal"))
                regimen_fiscal = _to_str(row.get("regimen_fiscal"))
                domicilio_fiscal_cp = _to_str(row.get("domicilio_fiscal_cp"))

                # si falta algo obligatorio, se salta y lo reporta
                if not rfc or not nombre_fiscal or not regimen_fiscal or not domicilio_fiscal_cp:
                    saltados += 1
                    continue

                data = {
                    "tipo_persona": _to_str(row.get("tipo_persona")) or Cliente.TipoPersona.MORAL,
                    "rfc": rfc,
                    "nombre_fiscal": nombre_fiscal,
                    "regimen_fiscal": regimen_fiscal_a_json([regimen_fiscal]),
                    "domicilio_fiscal_cp": domicilio_fiscal_cp,

                    "uso_cfdi_default": _to_str(row.get("uso_cfdi_default")),
                    "email_cfdi": _to_str(row.get("email_cfdi")),
                    "nombre_comercial": _to_str(row.get("nombre_comercial")),
                    "telefono": _to_str(row.get("telefono")),

                    "calle": _to_str(row.get("calle")),
                    "num_ext": _to_str(row.get("num_ext")),
                    "num_int": _to_str(row.get("num_int")),
                    "colonia": _to_str(row.get("colonia")),
                    "localidad": _to_str(row.get("localidad")),
                    "municipio": _to_str(row.get("municipio")),
                    "estado": _to_str(row.get("estado")),
                    "pais": _to_str(row.get("pais")) or "México",
                    "cp": _to_str(row.get("cp")),
                    "referencias": _to_str(row.get("referencias")),

                    "forma_pago_default": _to_str(row.get("forma_pago_default")),
                    "metodo_pago_default": _to_str(row.get("metodo_pago_default")),
                    "activo": _to_bool_si_no(row.get("activo"), default=True),
                }

                try:
                    obj, created = Cliente.objects.update_or_create(
                        rfc=rfc,  # es unique【turn2file0†models.py†L34-L44】
                        defaults=data,
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                except Exception as e:
                    errores.append(f"Fila {idx} (RFC {rfc}): {e}")

            if errores:
                messages.warning(
                    request,
                    f"Importación terminada. Creados: {creados}, Actualizados: {actualizados}, Saltados: {saltados}. "
                    f"Errores: {len(errores)} (ver consola)."
                )
                for err in errores[:30]:
                    print(err)
            else:
                messages.success(
                    request,
                    f"Importación terminada. Creados: {creados}, Actualizados: {actualizados}, Saltados: {saltados}."
                )

        except Exception as e:
            messages.error(request, f"Error al importar: {e}")

        return redirect("importar_clientes")

    return render(request, "catalogos/importar_clientes.html")

# --- Fin Importaciones ---
